import os
import json
from openpyxl import Workbook
from openpyxl import load_workbook
import xlsxwriter  # Ensure you have installed this library: pip install xlsxwriter

def optimal(EditedFields, fields_to_fill, optimal_output_path, benchmark_path, model_name):   
    # Load the optimal output for comparison
    with open(optimal_output_path, "r", encoding="utf-8") as f:
        optimal_output = json.load(f)

    # Initialize benchmark metrics
    total_fields = len(fields_to_fill)
    correctly_filled = 0
    incorrectly_filled = 0
    blank_fields = 0  # Tracks correctly left blank fields
    error_log = []  # List to store error messages

    # Compare EditedFields with optimalOutput
    for field_name in fields_to_fill:
        optimal_value = optimal_output.get(field_name, None)
        edited_value = EditedFields.get(field_name, None)

        if isinstance(edited_value, str) and isinstance(optimal_value, str):
            if not edited_value.strip() and not optimal_value.strip():  # Correctly left blank
                blank_fields += 1
            elif edited_value.strip().lower() == optimal_value.strip().lower():
                correctly_filled += 1
            elif edited_value.strip():  # Field was filled but incorrectly
                incorrectly_filled += 1
                error_log.append(f"Field '{field_name}' filled incorrectly. Expected: '{optimal_value}', Got: '{edited_value}'")
            else:  # Field was left blank incorrectly
                incorrectly_filled += 1
                error_log.append(f"Field '{field_name}' left blank incorrectly. Expected: '{optimal_value}', Got: ''")
        elif edited_value is None and optimal_value is None:  # Correctly left blank
            blank_fields += 1
        elif edited_value == optimal_value:
            correctly_filled += 1
        elif edited_value is not None:  # Field was filled but incorrectly
            incorrectly_filled += 1
            error_log.append(f"Field '{field_name}' filled incorrectly. Expected: '{optimal_value}', Got: '{edited_value}'")
        else:  # Field was left blank incorrectly
            incorrectly_filled += 1
            error_log.append(f"Field '{field_name}' left blank incorrectly. Expected: '{optimal_value}', Got: ''")

    # Add fields filled incorrectly that were not in optimalOutput
    extra_fields = set(EditedFields.keys()) - set(optimal_output.keys())
    for extra_field in extra_fields:
        incorrectly_filled += 1
        error_log.append(f"Extra field '{extra_field}' filled. Value: '{EditedFields[extra_field]}'")

    # Write errors to log.txt
    log_path = os.path.join(os.path.dirname(benchmark_path), "log.txt")
    with open(log_path, "w", encoding="utf-8") as log_file:
        log_file.write("\n".join(error_log))

    # Open or create the XLSX file
    if os.path.exists(benchmark_path):
        workbook = load_workbook(benchmark_path)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        # Write headers if creating a new file
        worksheet.append(["Model", "Total Fields", "Correctly Filled", "Incorrectly Filled", "Blank Fields"])

    # Check if the model already exists in the file
    model_exists = False
    for row in worksheet.iter_rows(min_row=2, max_col=1):  # Removed values_only=True
        if row[0].value == model_name:  # Access the cell value
            model_exists = True
            row_index = row[0].row  # Access the row index from the cell object
            worksheet.cell(row=row_index, column=2, value=total_fields)
            worksheet.cell(row=row_index, column=3, value=correctly_filled)
            worksheet.cell(row=row_index, column=4, value=incorrectly_filled)
            worksheet.cell(row=row_index, column=5, value=blank_fields)
            break

    # If the model does not exist, append a new row
    if not model_exists:
        worksheet.append([model_name, total_fields, correctly_filled, incorrectly_filled, blank_fields])

    # Save the workbook
    workbook.save(benchmark_path)

    print(f"\nBenchmark results updated in:\n{os.path.abspath(benchmark_path)}")